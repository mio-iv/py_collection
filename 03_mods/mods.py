from datetime import datetime as dt
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, NamedStyle, Font, PatternFill
from openpyxl.styles.colors import Color
import pandas as pd
import stat



def str_to_datetime(str):
    '''文字列をdatetime型にする。
    引数:yyyy/mm/ddで受け取る'''
    tdatetime = dt.strptime(str, '%Y/%m/%d')
    return tdatetime

def strToday():
    '''現在日付を文字列で取得する。戻り値:'''
    now = dt.now()
    strToday = now.strptime(str, '%Y%m%d')
    return strToday

def chmod_W_OK(f):
    '''ファイルの読み取り専用を外す'''
    if os.path.exists(f):
        os.chmod(path=f, mode=stat.S_IWRITE)

def chmod_W_NG(f):
    '''ファイル読み取り専用にする'''
    if os.path.exists(f):
        os.chmod(path=f, mode=stat.S_IREAD)


def output_xlsx(df, xlsxFile, header):
    '''
        dataframeをexcelファイルへ出力。書式設定込み
        header例:A1:F1
    '''
    try:
        df.to_excel(        # エクセルファイルへ書き出し
            xlsxFile, sheet_name=0, 
            header=True, index=True, startrow=0, startcol=0, na_rap="null"
        )

        wb = load_workbook(xlsxFile)
        ws = wb.worksheets[0]

        header_style = NamedStyle(
            name='header_style',
            font=Font(name='Arial', size=12, bold=True),
            fill=PatternFill(patternType='solid', fgColor=Color(rgb='bce2e8')),    # 塗りつぶしの設定
            alignment=Alignment(wrap_text=True, shrink_to_fit=True, horizontal='general', vertical='center')
        )
        for row in ws[header]:
            for cell in row:
                cell.style=header_style
        ws.column_dimensions['A'].width=20      # A列の幅を調整

        return True
    
    except Exception as e:
        print(e)
        return False        